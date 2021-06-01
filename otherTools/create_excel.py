# -*- coding:utf-8 -*-
from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime
import time
import os
from wuliu.settings import BASE_DIR


def create_excel(excel_name,data,result_zi):
	# 创建excel
	excel_name = excel_name
	filename = excel_name + datetime.now().strftime('%Y-%m-%d %H-%M')
	# BASE_DIR D:\Work\112.126.62.178\wuliu
	path_gen = BASE_DIR + '/'
	path_file = 'staticfiles/excel/' + filename + '.xlsx'
	path = path_gen + path_file
	print('path_file:',path_file)
	# print('path:',path)

	workbook = Workbook()

	# 重命名sheet名字
	workbook[workbook.sheetnames[0]].title = excel_name

	# 启动sheet
	sheet = workbook.active

	data = data

	# 找出题目
	title_lis = []
	for i in data[0].keys():
		title_lis.append(i)

	# 根据题目找注释
	title_chinese = []
	for i in title_lis:
		for k in result_zi:
			if i == k[0]:
				title_chinese.append(k[1])

	# 写入题目
	sheet.append(title_chinese)

	# 找出内容
	content_lis = []
	content_lis_all = []
	for i in data:
		content_lis.append(i.values())

	for i in content_lis:
		content_lis_all.append(list(i))

	# 写入内容
	for row in content_lis_all:
		sheet.append(row)

	workbook.save(filename=path)
	print('表格生成完成')
	return path_file


# 数据
data = [{
			"id": 4,
			"userId": 128,
			"userName": "\u90ed\u6bd3\u5251",
			"goodsName": "\u706b\u7bad",
			"money": 555567.0,
			"status": 1,
			"costMing": 2,
			"detail": "3",
			"createTime": "1589871856680",
			"wuliuId": 2,
			"tixianId": 'null'
		}, {
			"id": 3,
			"userId": 128,
			"userName": "\u90ed\u6bd3\u5251",
			"goodsName": "\u706b\u7bad",
			"money": 5555666.0,
			"status": 1,
			"costMing": 1,
			"detail": "3",
			"createTime": "1589871848101",
			"wuliuId": 2,
			"tixianId": 'null'
		}, {
			"id": 2,
			"userId": 128,
			"userName": "\u90ed\u6bd3\u5251",
			"goodsName": "\u706b\u7bad",
			"money": 11.0,
			"status": 1,
			"costMing": 1,
			"detail": "3",
			"createTime": "1589784484207",
			"wuliuId": 1,
			"tixianId": 'null'
		}]
result_zi = (('id', '流水id'), ('userId', '用户ID'), ('userName', '用户名'), ('goodsName', '托运物品'), ('money', '金额'), ('status', '收支状态（0：支出，1：收入）'), ('costMing', '费用说明（1：运费，2：押金，3：提现）'), ('detail', '明细（1：银行卡支付，2：微信支付，3：余额支付，4：提现银行卡账户名）'), ('createTime', '收支时间'), ('wuliuId', '物流订单id'), ('tixianId', '提现申请id'), ('shanchu', '0：显示，1：伪删除'))


# create_excel('信用管理',data=data,result_zi=result_zi)

